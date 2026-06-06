"""
Offline unit tests for the NBA injury-clip pipeline.

These exercise the *pure* logic (candidate detection, name matching, the
Hashtag Basketball injury-page parsing/confirm, schedule mapping, clip-URL
building, xlsx output) with synthetic data, so they run anywhere -- no
network, no stats.nba.com / hashtagbasketball.com access required.

    pip install -r requirements.txt pytest
    pytest test_nba_injury_clips.py -v
"""

import datetime as dt

import numpy as np
import pandas as pd

import nba_injury_clips as nic
import nba_injury_clips_batch as batch
import hashtag_injuries as hi


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
PBP_COLS = [
    "EVENTNUM", "EVENTMSGTYPE", "PLAYER1_ID", "PLAYER1_NAME", "PLAYER2_ID",
    "PERIOD", "PCTIMESTRING", "HOMEDESCRIPTION", "VISITORDESCRIPTION",
    "NEUTRALDESCRIPTION",
]


def _pbp(rows):
    """Build a play-by-play frame mimicking get_play_by_play()'s output."""
    df = pd.DataFrame(rows, columns=PBP_COLS)
    return df.sort_values("EVENTNUM").reset_index(drop=True)


# ---------------------------------------------------------------------------
# find_injury_candidates
# ---------------------------------------------------------------------------
def test_find_injury_candidates_basic():
    # Alpha: out@2 then back in@4   -> NOT a candidate (returned)
    # Bravo: in@2 then out@4        -> candidate, injury play = evt3
    # Charlie: out@6, never returns -> candidate, injury play = evt5
    # Delta: in@6                   -> NOT a candidate (came in, never out)
    SUB = nic.SUBSTITUTION
    pbp = _pbp([
        (1, 1, 0, "", 0, 1, "11:00", "Smith 2PT", None, None),
        (2, SUB, 10, "Alpha", 20, 1, "10:30", None, None, None),
        (3, 1, 0, "", 0, 1, "09:00", None, "Bravo layup", None),
        (4, SUB, 20, "Bravo", 10, 2, "05:00", None, None, None),
        (5, 1, 0, "", 0, 2, "04:00", "Charlie dunk", None, None),
        (6, SUB, 30, "Charlie", 40, 3, "02:00", None, None, None),
    ])
    cands = {c["player"]: c for c in nic.find_injury_candidates(pbp)}

    assert set(cands) == {"Bravo", "Charlie"}

    bravo = cands["Bravo"]
    assert bravo["period"] == 2
    assert bravo["clock"] == "05:00"
    assert bravo["injury_play_eventnum"] == 3
    assert bravo["injury_play_desc"] == "Bravo layup"

    charlie = cands["Charlie"]
    assert charlie["period"] == 3
    assert charlie["injury_play_eventnum"] == 5
    assert charlie["injury_play_desc"] == "Charlie dunk"


def test_find_injury_candidates_nan_description():
    # The play before the sub-out has all-NaN descriptions. float('nan') is
    # truthy, so a naive `a or b or c` chain would wrongly return NaN; we must
    # get the "(no description)" fallback instead.
    SUB = nic.SUBSTITUTION
    pbp = _pbp([
        (1, 1, 0, "", 0, 1, "11:00", np.nan, np.nan, np.nan),
        (2, SUB, 50, "Echo", 60, 1, "10:00", np.nan, np.nan, np.nan),
    ])
    cands = nic.find_injury_candidates(pbp)
    assert len(cands) == 1
    assert cands[0]["player"] == "Echo"
    assert cands[0]["injury_play_desc"] == "(no description)"


def test_find_injury_candidates_none_when_all_return():
    SUB = nic.SUBSTITUTION
    pbp = _pbp([
        (1, SUB, 10, "Alpha", 20, 1, "10:00", None, None, None),
        (2, SUB, 20, "Bravo", 10, 1, "08:00", None, None, None),
    ])
    # Alpha out@1 then in@2; Bravo in@1 then out@2 -> Bravo is the candidate.
    cands = nic.find_injury_candidates(pbp)
    assert [c["player"] for c in cands] == ["Bravo"]


# ---------------------------------------------------------------------------
# _normalize_name
# ---------------------------------------------------------------------------
def test_normalize_name():
    assert nic._normalize_name("• Jaren Jackson Jr.") == "jaren jackson"
    assert nic._normalize_name("D'Angelo Russell") == "d angelo russell"
    assert nic._normalize_name("Gary Payton II") == "gary payton"
    assert nic._normalize_name(None) == ""
    assert nic._normalize_name(123) == ""


# ---------------------------------------------------------------------------
# hashtag_injuries: detail-page parsing
# ---------------------------------------------------------------------------
INJURY_PAGE_HTML = """
<html><body>
<h1><span id="ContentPlaceHolder1_FormView1_NOTESLabel">Sprained left ankle</span></h1>
<table id="ContentPlaceHolder1_GridView1">
  <tr><th scope="col">PLAYER</th><th scope="col">TEAM</th>
      <th scope="col">INJURED ON</th><th scope="col">RETURNED</th>
      <th scope="col">DAYS MISSED</th></tr>
  <tr><td>Donovan Mitchell</td><td>Cavaliers</td>
      <td><span id="x_Label1_0">07 April 2025</span></td>
      <td><span id="x_Label2_0">19 April 2025</span></td><td>12</td></tr>
  <tr><td>Still Out</td><td>Heat</td>
      <td><span id="x_Label1_1">01 May 2025</span></td>
      <td></td><td></td></tr>
</table>
</body></html>
"""


def test_parse_injury_page():
    df = hi.parse_injury_page(INJURY_PAGE_HTML)
    assert list(df.columns) == [
        "player", "player_norm", "team", "date", "returned",
        "days_missed", "injury_desc",
    ]
    assert len(df) == 2

    dm = df.iloc[0]
    assert dm["player"] == "Donovan Mitchell"
    assert dm["player_norm"] == "donovan mitchell"
    assert dm["team"] == "Cavaliers"
    assert dm["date"] == dt.date(2025, 4, 7)
    assert dm["returned"] == dt.date(2025, 4, 19)
    assert dm["days_missed"] == 12
    assert dm["injury_desc"] == "Sprained left ankle"

    out = df.iloc[1]
    assert out["date"] == dt.date(2025, 5, 1)
    assert out["returned"] is None          # blank RETURNED
    assert hi._missing(out["days_missed"])   # blank DAYS MISSED


def test_parse_injury_page_missing_table():
    df = hi.parse_injury_page("<html><body>no table here</body></html>")
    assert df.empty
    assert "player" in df.columns


def test_parse_slug_results():
    html = """
      <table id="ContentPlaceHolder1_GridView2">
        <tr><th>INJURY</th></tr>
        <tr><td><a href="/injury/sprained-left-ankle">Sprained left ankle</a></td></tr>
        <tr><td><a href="/injury/sore-right-knee">Sore right knee</a></td></tr>
      </table>
    """
    slugs = hi.parse_slug_results(html)
    assert slugs == {
        "sprained-left-ankle": "Sprained left ankle",
        "sore-right-knee": "Sore right knee",
    }


def test_hidden_fields():
    html = (
        '<input type="hidden" name="__VIEWSTATE" id="__VIEWSTATE" value="VS123" />'
        '<input type="hidden" name="__VIEWSTATEGENERATOR" '
        'id="__VIEWSTATEGENERATOR" value="GEN9" />'
        '<input type="hidden" name="__EVENTVALIDATION" '
        'id="__EVENTVALIDATION" value="EV456" />'
    )
    f = hi._hidden_fields(html)
    assert f == {
        "__VIEWSTATE": "VS123",
        "__VIEWSTATEGENERATOR": "GEN9",
        "__EVENTVALIDATION": "EV456",
    }


# ---------------------------------------------------------------------------
# hashtag_injuries: severity + confirm
# ---------------------------------------------------------------------------
def test_format_severity():
    assert hi.format_severity(14, dt.date(2024, 1, 15)) == "14 days"
    assert hi.format_severity(12.0, None) == "12 days"          # float -> int
    assert hi.format_severity(None, None) == "unknown (no return logged)"
    assert hi.format_severity(np.nan, None) == "unknown (no return logged)"
    assert hi.format_severity(None, dt.date(2024, 1, 15)) == "unknown"


def _injuries(rows):
    return pd.DataFrame(rows, columns=[
        "player", "player_norm", "team", "date", "returned",
        "days_missed", "injury_desc",
    ])


def test_confirm_injury_in_window():
    df = _injuries([
        ["LeBron James", "lebron james", "Lakers", dt.date(2023, 10, 25),
         dt.date(2023, 11, 8), 14, "Sore left foot"],
    ])
    note = hi.confirm_injury("LeBron James", dt.date(2023, 10, 24), df, 2)
    assert note == "Sore left foot (14 days)"


def test_confirm_injury_lastname_fallback_and_still_out():
    df = _injuries([
        ["Bradley Beal", "bradley beal", "Suns", dt.date(2023, 10, 24),
         None, np.nan, "Back spasms"],
    ])
    note = hi.confirm_injury("B. Beal", dt.date(2023, 10, 24), df, 2)
    assert note == "Back spasms (unknown (no return logged))"


def test_confirm_injury_out_of_window_and_empty():
    df = _injuries([
        ["LeBron James", "lebron james", "Lakers", dt.date(2023, 11, 20),
         None, np.nan, "foot"],
    ])
    assert hi.confirm_injury("LeBron James", dt.date(2023, 10, 24), df, 2) is None

    empty = _injuries([]).iloc[0:0]
    assert hi.confirm_injury("Anyone", dt.date(2023, 10, 24), empty, 2) is None


def test_confirm_injury_skips_null_dates():
    # A null/NaN date must not crash the scan (iterrows would upcast it).
    df = _injuries([
        ["Ghost Player", "ghost player", "Nets", None, None, np.nan, "n/a"],
        ["LeBron James", "lebron james", "Lakers", dt.date(2023, 10, 25),
         dt.date(2023, 11, 8), 14, "foot"],
    ])
    assert hi.confirm_injury("LeBron James", dt.date(2023, 10, 24), df, 2) \
        == "foot (14 days)"


# ---------------------------------------------------------------------------
# batch: schedule mapping
# ---------------------------------------------------------------------------
def test_match_team():
    team_games = {"Los Angeles Lakers": [], "Boston Celtics": []}
    assert batch.match_team("Lakers", team_games) == "Los Angeles Lakers"
    assert batch.match_team("Celtics", team_games) == "Boston Celtics"
    assert batch.match_team("Nobody", team_games) is None
    assert batch.match_team("", team_games) is None


def test_find_injury_game_picks_closest_on_or_before():
    team_games = {
        "Los Angeles Lakers": [
            (dt.date(2024, 1, 1), "G1"),
            (dt.date(2024, 1, 3), "G2"),   # closest on/before log date
            (dt.date(2024, 1, 6), "G3"),   # after the log date -> excluded
        ]
    }
    assert batch.find_injury_game("Lakers", dt.date(2024, 1, 4), team_games) \
        == (dt.date(2024, 1, 3), "G2")


def test_find_injury_game_no_game_in_window():
    team_games = {"Los Angeles Lakers": [(dt.date(2024, 1, 1), "G1")]}
    assert batch.find_injury_game("Lakers", dt.date(2024, 1, 10), team_games) is None


def test_names_match():
    assert batch.names_match("lebron james", "lebron james")
    assert batch.names_match("l james", "lebron james")    # initial + last name
    assert not batch.names_match("kevin durant", "kevin love")
    assert not batch.names_match("", "lebron james")


# ---------------------------------------------------------------------------
# extract_clip_url (videoevents monkeypatched -- no network)
# ---------------------------------------------------------------------------
class _FakeVideoEvents:
    payload = {}

    def __init__(self, *a, **k):
        pass

    def get_dict(self):
        return type(self).payload


def _set_video_payload(monkeypatch, entry):
    _FakeVideoEvents.payload = {"resultSets": {"Meta": {"videoUrls": [entry]}}}
    monkeypatch.setattr(nic.videoevents, "VideoEvents", _FakeVideoEvents)


def test_extract_clip_url_direct_url(monkeypatch):
    _set_video_payload(monkeypatch, {"lurl": "https://videos.nba.com/x/large.mp4"})
    url = nic.extract_clip_url("0022300001", 42, (2023, 10, 25))
    assert url == "https://videos.nba.com/x/large.mp4"


def test_extract_clip_url_builds_from_uuid(monkeypatch):
    _set_video_payload(monkeypatch, {"uuid": "abc123"})
    url = nic.extract_clip_url("0022300001", 42, (2023, 10, 5))
    assert url == (
        "https://videos.nba.com/nba/pbp/media/2023/10/05/"
        "0022300001/42/abc123_1280x720.mp4"
    )


def test_extract_clip_url_none_when_missing(monkeypatch):
    _set_video_payload(monkeypatch, {})           # no url, no uuid
    assert nic.extract_clip_url("0022300001", 42, (2023, 10, 5)) is None


# ---------------------------------------------------------------------------
# write_xlsx round-trip
# ---------------------------------------------------------------------------
def test_write_xlsx_roundtrip(tmp_path):
    from openpyxl import load_workbook

    df = pd.DataFrame(
        [["LeBron James", "2023-10-25", "sore foot", "14 days", "http://clip"]],
        columns=batch.COLUMNS,
    )
    out = tmp_path / "out.xlsx"
    batch.write_xlsx(df, str(out))

    wb = load_workbook(out)
    ws = wb["Injuries"]
    assert [c.value for c in ws[1]] == batch.COLUMNS
    assert [c.value for c in ws[2]] == [
        "LeBron James", "2023-10-25", "sore foot", "14 days", "http://clip"
    ]
    assert ws.freeze_panes == "A2"
