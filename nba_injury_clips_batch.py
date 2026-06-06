"""
nba_injury_clips_batch.py
=========================

INJURY-LOG-DRIVEN driver. Instead of scanning every game, it starts from the
Hashtag Basketball NBA Injury Database (see hashtag_injuries.py) and only
processes the games that correspond to a logged injury.

For each logged injury it:
  1. finds the game that player's team played in the injury window,
  2. locates the play where the player exited and didn't return,
  3. resolves the clip URL for that play.

Output spreadsheet columns:
    name | date | injury_desc | severity | play_by_play_url

  - date            = the actual game date the injury occurred
  - injury_desc     = the injury-type description from the log
  - severity        = TIME MISSED, taken straight from the log's precomputed
                      "days missed" (e.g. "14 days"), or
                      "unknown (no return logged)" when the player is still out
  - play_by_play_url= clip of the injury play (blank if no in-game exit found)

Requires nba_injury_clips.py and hashtag_injuries.py in the same folder.

SETUP
-----
    pip install nba_api requests pandas openpyxl lxml

USAGE
-----
    python nba_injury_clips_batch.py --season 2023-24
    python nba_injury_clips_batch.py --season 2024-25
    python nba_injury_clips_batch.py --season 2025-26
    python nba_injury_clips_batch.py              # all three seasons
    python nba_injury_clips_batch.py --finalize   # merge checkpoint -> .xlsx

The injury log is scraped once and cached on disk (.hashtag_cache/), so the
first run is slower. nba_api still hits an unofficial endpoint, so this sleeps
between calls and checkpoints every injury. If it stops or you get blocked,
just re-run -- it resumes.
"""

import sys
import csv
import os
import time
from datetime import datetime, timedelta
from collections import defaultdict

import pandas as pd
from nba_api.stats.endpoints import leaguegamelog

from nba_injury_clips import (
    get_play_by_play,
    find_injury_candidates,
    extract_clip_url,
    _normalize_name,
    SLEEP_BETWEEN_CALLS,
    INJURY_WINDOW_DAYS,
)
from hashtag_injuries import fetch_injury_database, format_severity

SEASONS = ["2023-24", "2024-25", "2025-26"]
SEASON_TYPES = ["Regular Season", "Playoffs"]

RESULTS_CSV = "nba_injuries_2023_2026.csv"
PROGRESS_FILE = "processed_injuries.txt"
OUTPUT_XLSX = "nba_injuries_2023_2026.xlsx"

COLUMNS = ["name", "date", "injury_desc", "severity", "play_by_play_url"]


# ---------------------------------------------------------------------------
# Schedule, so we can map (team, date) -> game_id
# ---------------------------------------------------------------------------
def build_schedule(season):
    team_games = defaultdict(list)  # TEAM_NAME -> [(date, game_id)]
    for stype in SEASON_TYPES:
        log = leaguegamelog.LeagueGameLog(
            season=season, season_type_all_star=stype
        ).get_data_frames()[0]
        for _, r in log.iterrows():
            try:
                d = datetime.strptime(str(r["GAME_DATE"])[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
            team_games[r["TEAM_NAME"]].append((d, r["GAME_ID"]))
        time.sleep(SLEEP_BETWEEN_CALLS)
    for t in team_games:
        team_games[t].sort()
    return team_games


def match_team(team_nickname, team_games):
    """The log uses nicknames ('Lakers'); match into full TEAM_NAME."""
    nick = (team_nickname or "").lower().strip()
    if not nick:
        return None
    for name in team_games:
        if nick in name.lower():
            return name
    return None


def find_injury_game(team_nickname, injury_date, team_games):
    """The team's game on/just before the log date -> (date, game_id)."""
    name = match_team(team_nickname, team_games)
    if not name:
        return None
    lo = injury_date - timedelta(days=INJURY_WINDOW_DAYS)
    games = [(d, g) for (d, g) in team_games[name] if lo <= d <= injury_date]
    if not games:
        return None
    games.sort()
    return games[-1]  # closest game on/before the log date


def names_match(a_norm, b_norm):
    if a_norm == b_norm:
        return True
    aa, bb = a_norm.split(), b_norm.split()
    return bool(aa and bb and aa[-1] == bb[-1] and aa[0][:1] == bb[0][:1])


# ---------------------------------------------------------------------------
# Process one logged injury
# ---------------------------------------------------------------------------
def process_injury(injury, team_games, pbp_cache):
    """`injury` is a row (dict-like) from the Hashtag Basketball injury log."""
    severity = format_severity(injury["days_missed"], injury["returned"])
    url = ""
    used_date = injury["date"]

    match = find_injury_game(injury["team"], injury["date"], team_games)
    if match:
        gdate, gid = match
        used_date = gdate
        # Clip resolution is best-effort: if play-by-play / video can't be
        # fetched (e.g. stats.nba.com is soft-blocking this IP), still record
        # the injury -- just leave the clip URL blank instead of dropping it.
        if gid not in pbp_cache:
            try:
                pbp_cache[gid] = get_play_by_play(gid)
            except Exception:  # noqa: BLE001
                pbp_cache[gid] = None
            time.sleep(SLEEP_BETWEEN_CALLS)
        pbp = pbp_cache[gid]
        if pbp is not None:
            for c in find_injury_candidates(pbp):
                if names_match(_normalize_name(c["player"]), injury["player_norm"]):
                    ymd = (gdate.year, gdate.month, gdate.day)
                    time.sleep(SLEEP_BETWEEN_CALLS)
                    try:
                        url = extract_clip_url(gid, c["injury_play_eventnum"], ymd) or ""
                    except Exception:  # noqa: BLE001
                        url = ""
                    break

    return {
        "name": injury["player"],
        "date": used_date.isoformat(),
        "injury_desc": injury["injury_desc"],
        "severity": severity,
        "play_by_play_url": url,
    }


# ---------------------------------------------------------------------------
# Checkpointing
# ---------------------------------------------------------------------------
def load_processed():
    if not os.path.exists(PROGRESS_FILE):
        return set()
    with open(PROGRESS_FILE) as f:
        return {line.strip() for line in f if line.strip()}


def mark_processed(key):
    with open(PROGRESS_FILE, "a") as f:
        f.write(key + "\n")


def append_rows(rows):
    new = not os.path.exists(RESULTS_CSV)
    with open(RESULTS_CSV, "a", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        if new:
            w.writeheader()
        w.writerows(rows)


# ---------------------------------------------------------------------------
# Final spreadsheet
# ---------------------------------------------------------------------------
def finalize():
    if not os.path.exists(RESULTS_CSV):
        print(f"No checkpoint {RESULTS_CSV} -- nothing to finalize.")
        return
    df = pd.read_csv(RESULTS_CSV, dtype=str).fillna("")
    df = df.drop_duplicates().sort_values(["date", "name"]).reset_index(drop=True)
    write_xlsx(df, OUTPUT_XLSX)
    print(f"Wrote {len(df)} rows -> {OUTPUT_XLSX}")


def write_xlsx(df, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Injuries"

    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="left")

    for _, r in df.iterrows():
        ws.append([r[c] for c in headers])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    widths = {"name": 22, "date": 12, "injury_desc": 55,
              "severity": 22, "play_by_play_url": 70}
    for i, col in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            widths.get(col, 18)

    ws.freeze_panes = "A2"
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    args = sys.argv[1:]

    if "--finalize" in args:
        finalize()
        return

    seasons = [args[args.index("--season") + 1]] if "--season" in args else SEASONS
    processed = load_processed()

    print("Fetching the Hashtag Basketball injury log (cached on disk) ...")
    injuries = fetch_injury_database()

    for season in seasons:
        print(f"\n=== Season {season} ===")
        team_games = build_schedule(season)
        all_dates = [d for games in team_games.values() for (d, _) in games]
        if not all_dates:
            print("  no games found for this season; skipping.")
            continue
        lo, hi = min(all_dates), max(all_dates) + timedelta(days=INJURY_WINDOW_DAYS)

        season_injuries = injuries[
            injuries["date"].apply(lambda d: d is not None and lo <= d <= hi)
        ].reset_index(drop=True)
        print(f"{len(season_injuries)} logged injuries in {lo} .. {hi}.")

        pbp_cache = {}
        total = len(season_injuries)
        for i, inj in enumerate(season_injuries.to_dict("records"), 1):
            key = f"{inj['player_norm']}|{inj['date'].isoformat()}|{inj['injury_desc']}"
            if key in processed:
                continue
            try:
                row = process_injury(inj, team_games, pbp_cache)
                append_rows([row])
                mark_processed(key)
                processed.add(key)
                tag = "clip" if row["play_by_play_url"] else "no clip"
                print(f"  [{i}/{total}] {row['name']} {row['date']} "
                      f"| {row['severity']} | {tag}")
            except Exception as e:  # noqa: BLE001 - keep the run alive
                print(f"  [{i}/{total}] {inj['player']} FAILED: {e}")
            time.sleep(SLEEP_BETWEEN_CALLS)

    finalize()


if __name__ == "__main__":
    main()
